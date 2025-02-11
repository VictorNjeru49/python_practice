import openpyxl as xl 
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):

    wb = xl.load_workbook(filename) #'transactional.xlsx'
    try:
        sheet = wb['Sheet1']
        cell = sheet['A1']
        cell = sheet.cell(1, 1)

        print(cell.value)
        print(sheet.max_row)

        for row in range(2, sheet.max_row + 1):
            print(row)
            cell = sheet.cell(row, 3)  # Move this inside the loop
            corrected_value = cell.value * 0.9
            corrected_value_cell = sheet.cell(row, 4)
            corrected_value_cell.value = corrected_value

        # Corrected Reference parameters
        values = Reference(sheet,
                min_row=2,
                max_row=sheet.max_row,  # Changed to max_row
                min_col=4,
                max_col=4)
        
        chart = BarChart()
        chart.add_data(values)
        sheet.add_chart(chart, 'E2')

        wb.save(filename) #'transactional2.xlsx'
    except ModuleNotFoundError:
        print("The openpyxl module is not installed. Please install it using pip: pip install openpyxl")
